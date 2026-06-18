"""
EDU AI Agent Suite — ROI/TCO Calculator Generator
Generates EDU-AI-Suite-ROI-Calculator.xlsx for use in customer conversations.

Usage:
    pip install openpyxl
    python generate_roi_calculator.py

Output:
    EDU-AI-Suite-ROI-Calculator.xlsx (in the same directory as this script)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
GREEN_INPUT  = "D9F2D9"   # user-editable input cells
BLUE_CALC    = "D9E8F5"   # calculated / formula cells
ORANGE_ASSM  = "FFE8CC"   # assumption cells (editable but pre-filled)
YELLOW_HDR   = "FFD966"   # section headers
DARK_HDR     = "1F4E79"   # dark blue header rows (white text)
LIGHT_GREY   = "F2F2F2"   # alternating row fill
WHITE        = "FFFFFF"

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)

def _align(horizontal="left", wrap=False):
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)

def _border(style="thin"):
    s = Side(border_style=style, color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def _header_row(ws, row, values, col_start=1, bg=DARK_HDR, fg="FFFFFF", bold=True, size=11):
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=val)
        cell.fill = _fill(bg)
        cell.font = _font(bold=bold, size=size, color=fg)
        cell.alignment = _align("center")
        cell.border = _border()

def _input_cell(ws, row, col, value=None, fmt=None, label=None):
    """Green input cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(GREEN_INPUT)
    c.font = _font()
    c.alignment = _align()
    c.border = _border()
    if fmt:
        c.number_format = fmt
    return c

def _calc_cell(ws, row, col, formula=None, fmt=None):
    """Blue calculated cell."""
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = _fill(BLUE_CALC)
    c.font = _font(italic=True)
    c.alignment = _align("right")
    c.border = _border()
    if fmt:
        c.number_format = fmt
    return c

def _assm_cell(ws, row, col, value=None, fmt=None):
    """Orange assumption cell (pre-filled but editable)."""
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(ORANGE_ASSM)
    c.font = _font()
    c.alignment = _align()
    c.border = _border()
    if fmt:
        c.number_format = fmt
    return c

def _label(ws, row, col, text, bold=False, wrap=False, bg=LIGHT_GREY):
    c = ws.cell(row=row, column=col, value=text)
    if bg:
        c.fill = _fill(bg)
    c.font = _font(bold=bold)
    c.alignment = _align(wrap=wrap)
    c.border = _border()
    return c

def _section_header(ws, row, text, ncols=6, col_start=1, bg=YELLOW_HDR):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_start + ncols - 1
    )
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = _fill(bg)
    c.font = _font(bold=True, size=12)
    c.alignment = _align()
    c.border = _border()

# ===========================================================================
# Sheet 1 — Instructions
# ===========================================================================

def build_instructions(wb):
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_view.showGridLines = False

    _set_col_width(ws, "A", 5)
    _set_col_width(ws, "B", 55)
    _set_col_width(ws, "C", 30)

    # Title
    ws.merge_cells("B1:C1")
    c = ws["B1"]
    c.value = "EDU AI Agent Suite — ROI / TCO Calculator"
    c.fill = _fill(DARK_HDR)
    c.font = _font(bold=True, size=16, color="FFFFFF")
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "For use in customer conversations. All inputs are editable — the numbers are yours."
    c.font = _font(italic=True, size=10, color="595959")
    c.alignment = _align("center")

    rows = [
        (4, "HOW TO USE THIS CALCULATOR", True),
        (5, "1.  Start on the 'Institution Profile' tab. Enter your institution's basic profile "
            "— enrollment, staffing, current volumes, and costs. These drive the calculations "
            "on every other sheet.", False),
        (6, "2.  Go to 'Agent ROI'. Toggle each agent on or off (Y/N). Adjust the volume "
            "inputs and deflection-rate assumptions for each agent you plan to deploy. "
            "The labor savings and value calculations update automatically.", False),
        (7, "3.  Review 'AWS TCO'. Enter estimated monthly interaction volumes for each "
            "service tier. The sheet models Bedrock inference (Claude Sonnet and Haiku), "
            "AgentCore, DynamoDB, S3, Lambda, and supporting services.", False),
        (8, "4.  The 'Summary Dashboard' consolidates everything: total investment, total "
            "annual value, net benefit, ROI %, payback period, and a 3-year NPV. The bar "
            "chart shows cost vs. value by year.", False),
        (10, "COLOR LEGEND", True),
    ]

    for row, text, bold in rows:
        ws.merge_cells(f"B{row}:C{row}")
        c = ws.cell(row=row, column=2, value=text)
        c.font = _font(bold=bold, size=11 if not bold else 12)
        c.alignment = _align(wrap=True)
        ws.row_dimensions[row].height = 40 if not bold else 18

    legend = [
        (11, GREEN_INPUT, "Green cells",  "Input cells — enter your institution's numbers here"),
        (12, BLUE_CALC,   "Blue cells",   "Calculated cells — formula-driven; do not edit"),
        (13, ORANGE_ASSM, "Orange cells", "Assumption cells — pre-filled with reference data; adjust as needed"),
        (14, YELLOW_HDR,  "Yellow cells", "Section headers"),
    ]
    for row, color, label, desc in legend:
        ws.cell(row=row, column=2, value=label).fill = _fill(color)
        ws.cell(row=row, column=2).font = _font(bold=True)
        ws.cell(row=row, column=2).border = _border()
        ws.cell(row=row, column=3, value=desc).font = _font()
        ws.cell(row=row, column=3).border = _border()

    ws.merge_cells("B16:C16")
    c = ws["B16"]
    c.value = ("DISCLAIMER: All numbers in this calculator are illustrative. "
               "Actual outcomes depend on institution size, workflow volume, "
               "model selection, deployment configuration, and measured results "
               "from a properly baselined pilot. Reference outcomes (UA-Pulaski Tech, "
               "Highline College, Illinois Tech) are cited as peer benchmarks, not guarantees.")
    c.font = _font(italic=True, size=9, color="595959")
    c.alignment = _align(wrap=True)
    ws.row_dimensions[16].height = 60

# ===========================================================================
# Sheet 2 — Institution Profile
# ===========================================================================

def build_institution_profile(wb):
    ws = wb.create_sheet("Institution Profile", 1)
    ws.sheet_view.showGridLines = False

    _set_col_width(ws, "A", 4)
    _set_col_width(ws, "B", 42)
    _set_col_width(ws, "C", 20)
    _set_col_width(ws, "D", 20)

    _header_row(ws, 1, ["", "Institution Profile", "Your Value", "Notes"], col_start=1)
    ws.row_dimensions[1].height = 22

    _section_header(ws, 2, "INSTITUTION TYPE", ncols=4)

    # Institution type dropdown
    _label(ws, 3, 2, "Institution type")
    ic = _input_cell(ws, 3, 3, value="Community College")
    dv = DataValidation(
        type="list",
        formula1='"K-12 District,Community College,University,Online Program,Workforce Education"',
        allow_blank=False
    )
    ws.add_data_validation(dv)
    dv.add(ic)

    _label(ws, 4, 2, "Institution name (optional)")
    _input_cell(ws, 4, 3, value="[Your Institution]")

    _section_header(ws, 6, "STUDENT ENROLLMENT", ncols=4)
    profile_inputs = [
        (7,  "Total student enrollment (headcount)",         5000,   "#,##0"),
        (8,  "  — of which: under 18 (K-12 or dual-enroll)",500,    "#,##0"),
        (9,  "Peak enrollment inquiry volume (monthly)",     2000,   "#,##0"),
        (10, "FAFSA / financial aid inquiries (monthly)",    800,    "#,##0"),
    ]
    for row, label, default, fmt in profile_inputs:
        _label(ws, row, 2, label, bg=WHITE if row % 2 == 0 else LIGHT_GREY)
        _input_cell(ws, row, 3, value=default, fmt=fmt)

    _section_header(ws, 12, "STAFFING AND COST", ncols=4)
    staff_inputs = [
        (13, "Total admin / student services FTE",           12,    "#,##0"),
        (14, "Total advising / counseling FTE",              8,     "#,##0"),
        (15, "Total IT / operations FTE",                    6,     "#,##0"),
        (16, "Total faculty FTE",                            150,   "#,##0"),
        (17, "Avg fully-loaded staff cost ($/hour)",         45,    '"$"#,##0.00'),
        (18, "Avg fully-loaded faculty cost ($/hour)",       65,    '"$"#,##0.00'),
    ]
    for row, label, default, fmt in staff_inputs:
        _label(ws, row, 2, label, bg=WHITE if row % 2 == 0 else LIGHT_GREY)
        _input_cell(ws, row, 3, value=default, fmt=fmt)

    _section_header(ws, 20, "CURRENT WORKFLOW BASELINES", ncols=4)
    baseline_inputs = [
        (21, "Routine inquiry handling time (minutes/inquiry)", 8,    "#,##0.0"),
        (22, "Document processing cycle time (days, current)",  14,   "#,##0"),
        (23, "Students per advisor (current caseload)",         250,  "#,##0"),
        (24, "IT ticket mean time to resolve (hours)",          18,   "#,##0.0"),
        (25, "Grading hours per assignment (avg, faculty)",     0.5,  "#,##0.00"),
        (26, "% of inquiries received after business hours",    0.3,  "0%"),
    ]
    for row, label, default, fmt in baseline_inputs:
        _label(ws, row, 2, label, bg=WHITE if row % 2 == 0 else LIGHT_GREY)
        _input_cell(ws, row, 3, value=default, fmt=fmt)

    # Note box
    ws.merge_cells("B28:D28")
    c = ws["B28"]
    c.value = ("NOTE: These baseline numbers drive the ROI calculations on the 'Agent ROI' sheet. "
               "Capture your actual baseline data before the pilot — the pilot outcome is measured "
               "against these numbers.")
    c.font = _font(italic=True, size=9, color="595959")
    c.alignment = _align(wrap=True)
    ws.row_dimensions[28].height = 45

# ===========================================================================
# Sheet 3 — Agent ROI
# ===========================================================================

AGENTS = [
    ("01", "Student & Family Services Concierge",  "Service, Labor, Student journey"),
    ("02", "Personalized Tutor & Study Companion",  "Learning, Service"),
    ("03", "Educator Copilot",                     "Labor, Learning"),
    ("04", "Assessment, Grading & Feedback",        "Labor, Learning"),
    ("05", "Student Success & Proactive Engagement","Student journey, Labor, Risk"),
    ("06", "Pathway Navigator",                    "Student journey, Service"),
    ("07", "Document & Accessibility Services",    "Labor, Service, Student journey"),
    ("08", "Operations Service Desk",              "Labor, Service"),
]

def build_agent_roi(wb):
    ws = wb.create_sheet("Agent ROI", 2)
    ws.sheet_view.showGridLines = False

    _set_col_width(ws, "A", 4)
    _set_col_width(ws, "B", 35)
    _set_col_width(ws, "C", 12)
    _set_col_width(ws, "D", 18)
    _set_col_width(ws, "E", 18)
    _set_col_width(ws, "F", 18)
    _set_col_width(ws, "G", 18)
    _set_col_width(ws, "H", 20)

    _header_row(ws, 1,
        ["", "Agent", "Deploy?", "Monthly Volume", "Deflection Rate",
         "Hours Saved/Mo", "Value/Month ($)", "One-Time Build ($)"],
        col_start=1
    )
    ws.row_dimensions[1].height = 22

    # Reference cells from Institution Profile
    IP = "'Institution Profile'"

    ref_defaults = {
        "01": (2000, 0.65, 45, 35000, 95000),
        "02": (500,  0.55, 45, 12000, 75000),
        "03": (300,  0.60, 65, 15000, 70000),
        "04": (200,  0.50, 65, 10000, 65000),
        "05": (400,  0.45, 45, 10000, 80000),
        "06": (350,  0.50, 45, 10000, 70000),
        "07": (600,  0.70, 45, 18000, 75000),
        "08": (1200, 0.65, 45, 20000, 60000),
    }

    row = 2
    for agent_num, agent_name, roi_cats in AGENTS:
        bg = LIGHT_GREY if int(agent_num) % 2 == 0 else WHITE
        vol, defl, hourly, val_mo, build = ref_defaults[agent_num]

        _label(ws, row, 2, f"Agent {agent_num} — {agent_name}", bg=bg)
        _label(ws, row, 3, roi_cats, bg=bg, wrap=True)

        # Deploy toggle
        dvc = _input_cell(ws, row, 3, value="Y")
        dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(dvc)

        _input_cell(ws, row, 4, value=vol, fmt="#,##0")
        _assm_cell(ws, row, 5, value=defl, fmt="0%")

        # Hours saved = volume * deflection * handling_time(min)/60
        hrs_formula = f'=IF(C{row}="Y", D{row}*E{row}*({IP}!C21/60), 0)'
        _calc_cell(ws, row, 6, formula=hrs_formula, fmt="#,##0.0")

        # Value = hours * staff rate
        val_formula = f'=IF(C{row}="Y", F{row}*{IP}!C17, 0)'
        _calc_cell(ws, row, 7, formula=val_formula, fmt='"$"#,##0')

        _assm_cell(ws, row, 8, value=build, fmt='"$"#,##0')

        row += 1

    # Totals
    row += 1
    _section_header(ws, row, "TOTALS", ncols=8)
    row += 1
    _label(ws, row, 2, "Total hours saved / month (all active agents)", bold=True)
    _calc_cell(ws, row, 6, formula=f"=SUM(F2:F{row-2})", fmt="#,##0.0")
    _calc_cell(ws, row, 7, formula=f"=SUM(G2:G{row-2})", fmt='"$"#,##0')
    _calc_cell(ws, row, 8, formula=f"=SUM(H2:H{row-2})", fmt='"$"#,##0')

    row += 1
    _label(ws, row, 2, "Total annual value (labor savings)", bold=True)
    _calc_cell(ws, row, 7, formula=f"=G{row-1}*12", fmt='"$"#,##0')

    row += 1
    _label(ws, row, 2, "Total one-time build investment", bold=True)
    _calc_cell(ws, row, 8, formula=f"=H{row-2}", fmt='"$"#,##0')

    # Note on gateway build-once economics
    row += 2
    ws.merge_cells(f"B{row}:H{row}")
    c = ws.cell(row=row, column=2,
        value=("NOTE — Build-once economics: The MCP gateway, identity, audit trail, and HITL "
               "framework are built once on the first agent. Agents 2–8 inherit them — the one-time "
               "build cost above reflects this amortization. The 'One-Time Build' for Agent 01 is "
               "the heaviest; subsequent agents are primarily connector + agent-specific logic."))
    c.font = _font(italic=True, size=9, color="595959")
    c.alignment = _align(wrap=True)
    ws.row_dimensions[row].height = 50

# ===========================================================================
# Sheet 4 — AWS TCO
# ===========================================================================

def build_aws_tco(wb):
    ws = wb.create_sheet("AWS TCO", 3)
    ws.sheet_view.showGridLines = False

    _set_col_width(ws, "A", 4)
    _set_col_width(ws, "B", 38)
    _set_col_width(ws, "C", 18)  # unit
    _set_col_width(ws, "D", 18)  # volume/mo
    _set_col_width(ws, "E", 18)  # rate
    _set_col_width(ws, "F", 18)  # monthly cost

    _header_row(ws, 1,
        ["", "AWS Service / Cost Element", "Unit", "Monthly Volume", "Rate ($/unit)", "Monthly Cost ($)"],
        col_start=1
    )

    tco_rows = [
        # (section_header, [(label, unit, default_vol, default_rate)])
        ("BEDROCK INFERENCE", [
            ("Claude 3.5 Sonnet — input tokens",  "per 1K tokens", 5000,  0.003),
            ("Claude 3.5 Sonnet — output tokens", "per 1K tokens", 2500,  0.015),
            ("Claude 3 Haiku — input tokens",     "per 1K tokens", 10000, 0.00025),
            ("Claude 3 Haiku — output tokens",    "per 1K tokens", 4000,  0.00125),
            ("Bedrock Knowledge Bases — queries", "per query",     8000,  0.000004),
        ]),
        ("AGENTCORE / GATEWAY", [
            ("AgentCore — invocations",           "per invocation",1000000, 0.0000025),
            ("AgentCore Identity — tokens minted","per token",     500000,  0.000001),
            ("API Gateway (Option B fallback) — requests", "per M requests", 1, 3.50),
            ("Lambda — GB-seconds",               "per GB-sec",    500000,  0.0000166),
            ("Step Functions — state transitions","per 1K transitions", 500, 0.025),
        ]),
        ("STORAGE", [
            ("DynamoDB — on-demand reads (RRU)",  "per M RRU",     10,    0.125),
            ("DynamoDB — on-demand writes (WRU)", "per M WRU",     5,     0.625),
            ("DynamoDB — storage (GB)",           "GB/month",      5,     0.25),
            ("S3 — Standard storage (GB)",        "GB/month",      50,    0.023),
            ("S3 — PUT/COPY/POST requests (per K)","per K requests",100,  0.005),
            ("S3 Object Lock (audit WORM, GB)",   "GB/month",      10,    0.023),
        ]),
        ("COMPUTE AND NETWORKING", [
            ("ECS Fargate — vCPU-hours/month",    "vCPU-hours",    720,   0.04048),
            ("ECS Fargate — GB memory-hours/month","GB-hours",     1440,  0.004445),
            ("Application Load Balancer — LCUs/hr","LCU-hours",    720,   0.008),
            ("NAT Gateway — GB processed",        "GB",            100,   0.045),
            ("CloudWatch — custom metrics",       "per metric/mo", 20,    0.30),
            ("CloudWatch — log ingestion (GB)",   "GB",            5,     0.50),
        ]),
        ("SUPPORTING SERVICES", [
            ("KMS — API requests (per 10K)",      "per 10K calls", 100,   0.03),
            ("Secrets Manager — secrets/month",   "per secret",    10,    0.40),
            ("CloudTrail — events (per 100K)",    "per 100K events",50,   0.10),
            ("Textract (Agent 07) — pages",       "per page",      1000,  0.0015),
            ("Translate (Agent 07) — per char (M)","per M chars",  1,     15.00),
            ("SES — emails sent",                 "per 1K emails", 10,    0.10),
        ]),
    ]

    row = 2
    for section, items in tco_rows:
        _section_header(ws, row, section, ncols=6)
        row += 1
        for label, unit, vol, rate in items:
            bg = WHITE if row % 2 == 0 else LIGHT_GREY
            _label(ws, row, 2, label, bg=bg)
            _label(ws, row, 3, unit, bg=bg)
            _assm_cell(ws, row, 4, value=vol, fmt="#,##0")
            _assm_cell(ws, row, 5, value=rate, fmt='"$"#,##0.000000')
            _calc_cell(ws, row, 6, formula=f"=D{row}*E{row}", fmt='"$"#,##0.00')
            row += 1

    # Total
    row += 1
    _section_header(ws, row, "TOTAL MONTHLY AWS COST", ncols=6, bg=DARK_HDR)
    row += 1
    _label(ws, row, 2, "Total monthly AWS infrastructure cost", bold=True)
    _calc_cell(ws, row, 6, formula=f"=SUM(F2:F{row-2})", fmt='"$"#,##0.00')

    row += 1
    _label(ws, row, 2, "Total annual AWS infrastructure cost", bold=True)
    _calc_cell(ws, row, 6, formula=f"=F{row-1}*12", fmt='"$"#,##0')

    # Note
    row += 2
    ws.merge_cells(f"B{row}:F{row}")
    c = ws.cell(row=row, column=2,
        value=("NOTE: Rates above reflect approximate list prices as of mid-2025. "
               "Actual rates depend on region, EDP/Reserved pricing, committed use discounts, "
               "and AWS pricing changes. Verify current rates at aws.amazon.com/pricing before "
               "presenting to a customer. Bedrock inference is typically the largest variable "
               "line — model at peak enrollment volume, not steady-state."))
    c.font = _font(italic=True, size=9, color="595959")
    c.alignment = _align(wrap=True)
    ws.row_dimensions[row].height = 60

# ===========================================================================
# Sheet 5 — Summary Dashboard
# ===========================================================================

def build_summary_dashboard(wb):
    ws = wb.create_sheet("Summary Dashboard", 4)
    ws.sheet_view.showGridLines = False

    _set_col_width(ws, "A", 4)
    _set_col_width(ws, "B", 42)
    _set_col_width(ws, "C", 22)
    _set_col_width(ws, "D", 22)
    _set_col_width(ws, "E", 22)

    # Title
    ws.merge_cells("B1:E1")
    c = ws["B1"]
    c.value = "EDU AI Agent Suite — ROI Summary Dashboard"
    c.fill = _fill(DARK_HDR)
    c.font = _font(bold=True, size=16, color="FFFFFF")
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 32

    # Input: discount rate
    _section_header(ws, 3, "INPUTS", ncols=4, bg=ORANGE_ASSM)
    _label(ws, 4, 2, "Discount rate for NPV calculation")
    _assm_cell(ws, 4, 3, value=0.08, fmt="0%")

    # --- Investment summary ---
    _section_header(ws, 6, "INVESTMENT SUMMARY", ncols=4)
    inv_rows = [
        (7,  "Total one-time build investment (all agents)",
              "='Agent ROI'!H19",   '"$"#,##0'),
        (8,  "Annual AWS infrastructure run cost",
              "='AWS TCO'!F47",     '"$"#,##0'),
        (9,  "Annual managed service fee (SI — estimate)",
              None,                  '"$"#,##0'),
        (10, "Total annual run cost (AWS + managed service)",
              "=C8+C9",             '"$"#,##0'),
        (11, "Total 3-year total cost of ownership",
              "=C7+C10*3",          '"$"#,##0'),
    ]
    for row, label, formula, fmt in inv_rows:
        bg = LIGHT_GREY if row % 2 == 0 else WHITE
        _label(ws, row, 2, label, bg=bg)
        if formula:
            _calc_cell(ws, row, 3, formula=formula, fmt=fmt)
        else:
            _input_cell(ws, row, 3, value=120000, fmt=fmt)

    # --- Value summary ---
    _section_header(ws, 13, "VALUE SUMMARY", ncols=4)
    val_rows = [
        (14, "Total annual labor savings (all active agents)",
              "='Agent ROI'!G20",  '"$"#,##0'),
        (15, "Service improvement value (% of labor savings — editable)",
              None,                 '"$"#,##0'),
        (16, "Total annual value (labor + service)",
              "=C14+C15",          '"$"#,##0'),
    ]
    for row, label, formula, fmt in val_rows:
        bg = LIGHT_GREY if row % 2 == 0 else WHITE
        _label(ws, row, 2, label, bg=bg)
        if formula:
            _calc_cell(ws, row, 3, formula=formula, fmt=fmt)
        else:
            _assm_cell(ws, row, 3, value=50000, fmt=fmt)

    # --- ROI metrics ---
    _section_header(ws, 18, "ROI METRICS", ncols=4, bg=DARK_HDR)
    roi_rows = [
        (19, "Net annual benefit (Value − Annual run cost)",  "=C16-C10",  '"$"#,##0'),
        (20, "First-year ROI %",                              "=C16/C7",   "0%"),
        (21, "Steady-state ROI % (year 2+)",                  "=C16/C10",  "0%"),
        (22, "Payback period (months)",                       "=C7/MAX(C16-C10,1)*12", "#,##0.0"),
        (23, "3-year NPV",
              "=C16/(1+C4)^1 + C16/(1+C4)^2 + C16/(1+C4)^3 - C7 - C10/(1+C4)^1 - C10/(1+C4)^2 - C10/(1+C4)^3",
              '"$"#,##0'),
    ]
    for row, label, formula, fmt in roi_rows:
        bg = LIGHT_GREY if row % 2 == 0 else WHITE
        _label(ws, row, 2, label, bold=True, bg=bg)
        _calc_cell(ws, row, 3, formula=formula, fmt=fmt)

    # --- Year-by-year table for chart ---
    _section_header(ws, 25, "3-YEAR VIEW (for chart)", ncols=4, bg=YELLOW_HDR)
    _header_row(ws, 26, ["", "Category", "Year 1", "Year 2", "Year 3"], col_start=1,
                bg=DARK_HDR, fg="FFFFFF")

    year_rows = [
        (27, "Total Investment (build + run)",
              "=C7+C10",     "=C10",       "=C10"),
        (28, "Total Value (labor + service)",
              "=C16",        "=C16",       "=C16"),
        (29, "Net Benefit",
              "=C28-C27",    "=D28-D27",   "=E28-E27"),
        (30, "Cumulative Net Benefit",
              "=C29",        "=C29+D29",   "=C29+D29+E29"),
    ]
    for row, label, y1, y2, y3 in year_rows:
        bg = LIGHT_GREY if row % 2 == 0 else WHITE
        _label(ws, row, 2, label, bg=bg)
        for col, formula in [(3, y1), (4, y2), (5, y3)]:
            _calc_cell(ws, row, col, formula=formula, fmt='"$"#,##0')

    # Bar chart — Investment vs Value by year
    chart = BarChart()
    chart.type = "col"
    chart.title = "Annual Investment vs. Value"
    chart.y_axis.title = "USD"
    chart.x_axis.title = "Year"
    chart.style = 10
    chart.width = 20
    chart.height = 14

    data = Reference(ws, min_col=3, max_col=5, min_row=26, max_row=28)
    cats = Reference(ws, min_col=2, max_col=2, min_row=27, max_row=28)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "B32")

# ===========================================================================
# Main
# ===========================================================================

def main():
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    build_instructions(wb)
    build_institution_profile(wb)
    build_agent_roi(wb)
    build_aws_tco(wb)
    build_summary_dashboard(wb)

    # Set active sheet to Instructions
    wb.active = wb["Instructions"]

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "EDU-AI-Suite-ROI-Calculator.xlsx")
    wb.save(out_path)
    print(f"OK ROI calculator written to: {out_path}")
    print("   Open the file and start on the 'Instructions' tab.")
    print("   Green cells are inputs. Blue cells are calculated. Orange cells are assumptions.")

if __name__ == "__main__":
    main()
